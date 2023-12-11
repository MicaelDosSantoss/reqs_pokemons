import openpyxl 
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import json

def color_type(cell):
    if cell.value == 'grass':
        color =  PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'water':
        color =  PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'fire':
        color =  PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'electric':
        color =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'poison':
        color =  PatternFill(start_color="800080", end_color="800080", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'normal':
        color =  PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'dragon':
        color =  PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'fairy':
        color =  PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'dark':
        color =  PatternFill(start_color="808080", end_color="808080", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'fighting':
        color =  PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'ground':
        color =  PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'psychic':
        color =  PatternFill(start_color="FF1493", end_color="FF1493", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'bug':
        color =  PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'rock':
        color =  PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'ghost':
        color =  PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'ice':
        color =  PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'steel':
        color =  PatternFill(start_color="E5E4E2", end_color="E5E4E2", fill_type="solid")  
        cell.fill = color
    elif cell.value == 'flying':
        color =  PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")  
        cell.fill = color


excel = openpyxl.Workbook() # planilha do excel



print(excel.sheetnames)

excel.create_sheet('Pokedex') # Criação da tabela de Dados

excel_pg = excel['Pokedex']
excel_pg.sheet_view.showGridLines = False

with open('Pokedex.json','r') as Poke:
    pokedex_json = json.load(Poke)
title = ['número do pokemon','Nome de pokemon','Tipos']

excel_pg.append(title)

cont_tab = 1

for poke in pokedex_json:
    row_data = ([poke["numero"], poke["pokemon"]])

    for tipo in poke['tipos']:
        row_data.append(tipo)

        
    
    excel_pg.append(row_data)

    cont_tab += 1

align_center = Alignment(horizontal='center',vertical='center') # Alinhar os itens nas células do excel
border_style = Side(style='thin') # Inserir bordar na tabela
font_bold = Font(bold=True)


# Aplicar formatação para colunas A B
for row in excel_pg.iter_rows(min_row=1, max_row=excel_pg.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.alignment = align_center
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        if cell.row == 1:  # Tornar títulos em negrito
            cell.font = font_bold



excel_pg.merge_cells('C1:D1')

colunaD = excel_pg['D']

cont = 0

for col in colunaD:
    cont += 1
    
    if col.value is None:
        excel_pg.merge_cells(f'C{cont}:D{cont}')

for cell in excel_pg['C']:
    color_type(cell)
for cell in excel_pg['D']:
    color_type(cell)
 


excel.save('Pokedex.xlsx') # Salvar alterações